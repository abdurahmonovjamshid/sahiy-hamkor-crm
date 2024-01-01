from datetime import datetime

import requests
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.db.models.signals import post_delete, post_save, pre_save
from django.dispatch import receiver
from django.http import HttpResponse
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from pytz import timezone

from conf import settings

from .models import (CuttingEvent, Product, ProductProduction,
                     ProductReProduction, Sales, SalesEvent, SalesEvent2,
                     Warehouse, Component)


@login_required
def component_export_excel(request):
    components = Component.objects.exclude(parent=None)

    # Create a new workbook
    workbook = Workbook()
    worksheet = workbook.active

    search_query = request.GET.get('q')
    if search_query:
        components = components.filter(Q(name__icontains=search_query))

    # Apply filters based on request parameters
    filters = request.GET.dict()
    filters.pop('q', None)
    filters.pop('o', None)
    filters.pop('p', None)
    filters.pop('all', None)

    if filters:
        components = components.filter(**filters)

    # Write headers
    headers = ['Bo\'lim', 'Nomi', 'Narxi', 'O\'lchov birligi', 'Umumiy']
    worksheet.append(headers)

    # Write data rows
    for component in components:
        row = [
            component.parent.title,
            component.title,
            str(component.price)+'$',
            component.measurement,
            component.total,
        ]
        worksheet.append(row)

    # Set the response headers for file download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=component.xlsx'

    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 20
    worksheet.column_dimensions['H'].width = 20

    # Save workbook to response
    workbook.save(response)

    return response


@login_required
def export_excel(request):
    products = Product.objects.all()

    # Create a new workbook
    workbook = Workbook()
    worksheet = workbook.active

    search_query = request.GET.get('q')
    if search_query:
        products = products.filter(Q(name__icontains=search_query))

    # Apply filters based on request parameters
    filters = request.GET.dict()
    filters.pop('q', None)
    filters.pop('o', None)
    filters.pop('p', None)
    filters.pop('all', None)

    if filters:
        products = products.filter(**filters)

    # Write headers
    headers = ['Nomi', 'Tan narxi', 'Sotuv narxi', 'Kesilmaganlar soni', 'Kesilganlar soni',
               'Sotilganlar soni', 'Mavjud tovar narxi', 'Sotilgan tovar narxi']
    worksheet.append(headers)

    # Write data rows
    for product in products:
        product_price = 0
        for productcomponent in product.productcomponent_set.all():
            product_price += productcomponent.quantity*productcomponent.component.price
        product_price = "{:,.1f}".format(product_price)+'$'

        row = [
            product.name,
            product_price,
            str(product.price)+'$',
            product.total_new,
            product.total_cut,
            product.total_sold,
            "{:,.1f}".format(
                (product.total_new + product.total_cut)*product.price),
            "{:,.1f}".format(product.total_sold_price)
        ]
        worksheet.append(row)

    # Set the response headers for file download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=products.xlsx'

    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 20
    worksheet.column_dimensions['H'].width = 20

    # Save workbook to response
    workbook.save(response)

    return response


@login_required
def export_warehouse_excel(request):
    queryset = Warehouse.objects.all()

    # Create a new workbook
    workbook = Workbook()
    worksheet = workbook.active

    search_query = request.GET.get('q')
    if search_query:
        queryset = queryset.filter(Q(name__icontains=search_query))

    filters = request.GET.dict()
    filters.pop('q', None)
    filters.pop('o', None)
    filters.pop('p', None)
    filters.pop('all', None)

    if filters:
        queryset = queryset.filter(**filters)

    # Write headers
    headers = ['Komponent', 'Miqdori',
               'Keltirilgan narxi', 'Keltirilgan vaqti', 'Xodim']
    worksheet.append(headers)

    # Write data rows
    for warehouse in queryset:
        row = [
            str(warehouse.component),
            warehouse.quantity,
            warehouse.price,
            warehouse.arrival_time,
            warehouse.user.username,

        ]
        worksheet.append(row)

    # Remove time zone information from arrival_time
    for column in worksheet.iter_cols(min_col=4, max_col=4, min_row=2):
        for cell in column:
            if cell.value:
                cell.value = cell.value.replace(tzinfo=None)

    # Set column alignment
    for column in worksheet.columns:
        column_letter = get_column_letter(column[0].column)
        alignment = Alignment(horizontal='left')
        worksheet.column_dimensions[column_letter].alignment = alignment

    # Set column width for Arrival Time
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 20
    worksheet.column_dimensions['H'].width = 20

    # Set the response headers for file download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=warehouse.xlsx'

    # Save workbook to response
    workbook.save(response)

    return response


@login_required
def export_production_excel(request):
    queryset = ProductProduction.objects.all()

    # Create a new workbook
    workbook = Workbook()
    worksheet = workbook.active

    search_query = request.GET.get('q')
    if search_query:
        queryset = queryset.filter(Q(series=search_query))

    filters = request.GET.dict()
    filters.pop('q', None)
    filters.pop('o', None)
    filters.pop('p', None)
    filters.pop('all', None)

    if filters:
        queryset = queryset.filter(**filters)

    # Write headers
    headers = ['Seriya', 'Produkt', 'Kesilmaganlar soni',
               'Xodim', 'Ishlab chiqarilish vaqti']
    worksheet.append(headers)

    # Write data rows
    for production in queryset:
        row = [
            production.series,
            str(production.product),
            production.quantity,
            production.user.username,
            production.date.replace(
                tzinfo=None),  # Remove timezone information
        ]
        worksheet.append(row)

    # Set column width for Arrival Time
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 30
    worksheet.column_dimensions['H'].width = 20

    # Set the response headers for file download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=production.xlsx'

    # Save workbook to response
    workbook.save(response)

    return response


@login_required
def export_reproduction_excel(request):
    queryset = ProductReProduction.objects.all()

    # Create a new workbook
    workbook = Workbook()
    worksheet = workbook.active

    search_query = request.GET.get('q')
    if search_query:
        queryset = queryset.filter(Q(user__username__icontains=search_query))

    filters = request.GET.dict()
    filters.pop('q', None)
    filters.pop('o', None)
    filters.pop('p', None)
    filters.pop('all', None)
    if filters:
        queryset = queryset.filter(**filters)

    # Write headers
    headers = ['Xodim', 'Umumiy kesilganlar',
               'Kesilgan mahsulotlar', 'Kesilgan vaqti']
    worksheet.append(headers)

    # Write data rows
    for reproduction in queryset:
        cutting_events = reproduction.cutting.all()
        events = ", ".join(str(str(cutting_event.quantity_cut) + ' ta ' +
                           cutting_event.product.name) for cutting_event in cutting_events)
        total_cut = 0
        for cuttingevent in reproduction.cutting.all():
            total_cut += cuttingevent.quantity_cut
        row = [
            reproduction.user.username,
            total_cut,
            events,
            reproduction.date.replace(
                tzinfo=None),  # Remove timezone information
        ]
        worksheet.append(row)

    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 50
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 20
    worksheet.column_dimensions['H'].width = 20
    # Set the response headers for file download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reproduction.xlsx'

    # Save workbook to response
    workbook.save(response)

    return response


@login_required
def export_sales_excel(request):
    queryset = Sales.objects.all()

    # Create a new workbook
    workbook = Workbook()
    worksheet = workbook.active

    search_query = request.GET.get('q')
    if search_query:
        queryset = queryset.filter(
            Q(buyer__icontains=search_query) | Q(seller__icontains=search_query))

    filters = request.GET.dict()
    filters.pop('q', None)
    filters.pop('o', None)
    filters.pop('all', None)
    filters.pop('p', None)
    if filters:
        queryset = queryset.filter(**filters)

    # Write headers
    headers = ['Haridor', 'Sotuvchi',
               'Kesilgan mahsulotlar', 'Kesilmagan mahsulotlar', 'Narx', 'Xodim', 'Sotilgan sana']
    worksheet.append(headers)

    # Write data rows
    for sale in queryset:
        sales_events = sale.selling_cut.all()
        cuts = ", ".join(str(
            sale_event)+f' ({sale_event.single_sold_price} dan)' for sale_event in sales_events)

        sales_event2s = sale.selling.all()
        non_cuts = ", ".join(str(
            sale_event2)+f' ({sale_event2.single_sold_price} dan)' for sale_event2 in sales_event2s)

        sales_events = sale.selling_cut.all()
        total_price = sum(event.total_sold_price for event in sales_events)

        sales_events2 = sale.selling.all()
        total_price2 = sum(event.total_sold_price for event in sales_events2)
        formatted_price = "{:,.1f}".format(total_price + total_price2)

        row = [
            sale.buyer,
            sale.seller,
            cuts,
            non_cuts,
            formatted_price,
            sale.user.username if sale.user else '-',
            sale.date.replace(
                tzinfo=None),  # Remove timezone information
        ]
        worksheet.append(row)

    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 50
    worksheet.column_dimensions['D'].width = 50
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 20
    worksheet.column_dimensions['H'].width = 20
    # Set the response headers for file download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=sales.xlsx'

    # Save workbook to response
    workbook.save(response)

    return response


@receiver(post_save, sender=ProductProduction)
def subtract_component_total(sender, instance, created, **kwargs):
    if created:
        product_components = instance.product.productcomponent_set.all()
        product = instance.product
        product.total_new += instance.quantity
        product.save()

        for product_component in product_components:
            total_quantity_by_component = product_component.quantity * \
                instance.quantity
            component = product_component.component
            component.total -= total_quantity_by_component

            if component.total < 0:
                component.total = 0

            component.save()

            try:
                for tg_id in settings.ADMINS:
                    if component.total <= component.notification_limit:
                        message_text = f'{component.title} limitdan kamayib ketdi!'
                        url = f'https://api.telegram.org/bot{settings.TOKEN}/sendMessage'

                        inline_keyboard = [[{
                            'text': 'Ko\'rish',
                            'url': f'{settings.HOSTNAME}'+reverse('admin:seh_1_component_changelist')
                        }]]

                        message = {
                            'chat_id': tg_id,
                            'text': message_text,
                            'reply_markup': {
                                'inline_keyboard': inline_keyboard
                            }
                        }

                        response = requests.post(url, json=message)
                        print(response.text)

            except Exception as e:
                print(e)


@receiver(post_delete, sender=ProductProduction)
def delete_component_total(sender, instance, **kwargs):
    product_components = instance.product.productcomponent_set.all()
    product = instance.product
    product.total_new -= instance.quantity
    product.save()

    for product_component in product_components:
        total_quantity_by_component = product_component.quantity * instance.quantity
        component = product_component.component
        component.total += total_quantity_by_component
        component.save()


@receiver(post_save, sender=Warehouse)
def add_component_total(sender, instance, created, **kwargs):
    if created:
        print("qo'shildi" * 88)
        component = instance.component
        component.total += instance.quantity
        component.save()


@receiver(pre_save, sender=Warehouse)
def update_component_total_on_edit(sender, instance, **kwargs):
    if instance.pk:
        # Warehouse object is being edited (not a new object)
        old_warehouse = Warehouse.objects.get(pk=instance.pk)
        component = instance.component
        component.total = component.total - old_warehouse.quantity + instance.quantity
        component.save()


@receiver(post_delete, sender=Warehouse)
def update_component_total_on_delete(sender, instance, **kwargs):
    component = instance.component
    component.total -= instance.quantity
    component.save()


@receiver(post_save, sender=CuttingEvent)
def curring_create(sender, instance, created, **kwargs):
    if created:
        product = instance.product
        product.total_new -= instance.quantity_cut
        product.total_cut += instance.quantity_cut
        product.save()


@receiver(post_delete, sender=CuttingEvent)
def cutting_delete(sender, instance, **kwargs):
    product = instance.product
    product.total_new += instance.quantity_cut
    product.total_cut -= instance.quantity_cut
    product.save()


@receiver(post_save, sender=SalesEvent)
def sales_create(sender, instance, created, **kwargs):
    if created:
        product = instance.product
        product.total_cut -= instance.quantity_sold
        product.total_sold += instance.quantity_sold
        product.total_sold_price += instance.total_sold_price
        product.save()


@receiver(post_delete, sender=SalesEvent)
def sales_delete(sender, instance, **kwargs):

    product = instance.product
    product.total_cut += instance.quantity_sold
    product.total_sold -= instance.quantity_sold
    product.total_sold_price -= instance.total_sold_price
    product.save()

    sales = instance.sales
    if sales.selling_cut.all().count() == 0 and sales.selling.all().count() == 0:
        sales.delete()


@receiver(post_save, sender=SalesEvent2)
def sales2_create(sender, instance, created, **kwargs):
    if created:
        product = instance.product
        product.total_new -= instance.quantity_sold
        product.total_sold += instance.quantity_sold
        product.total_sold_price += instance.total_sold_price
        product.save()


@receiver(post_delete, sender=SalesEvent2)
def sales2_delete(sender, instance, **kwargs):
    product = instance.product
    product.total_new += instance.quantity_sold
    product.total_sold -= instance.quantity_sold
    product.total_sold_price -= instance.total_sold_price
    product.save()

    sales = instance.sales
    if sales.selling_cut.all().count() == 0 and sales.selling.all().count() == 0:
        sales.delete()
