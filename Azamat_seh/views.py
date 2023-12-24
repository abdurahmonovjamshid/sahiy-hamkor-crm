from datetime import datetime

import requests
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.db.models.signals import post_delete, post_save, pre_save
from django.dispatch import receiver
from django.http import HttpResponse
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from pytz import timezone

from conf import settings

from .models import (Product, ProductProduction, Sales, SalesEvent, Warehouse)


@receiver(post_save, sender=ProductProduction)
def subtract_component_total(sender, instance, created, **kwargs):
    if created:
        product_components = instance.product.productcomponent_set.all()
        product = instance.product
        product.total_new += instance.quantity
        product.save()

        for product_component in product_components:
            total_quantity_by_component = product_component.quantity * \
                instance.quantity
            component = product_component.component
            component.total -= total_quantity_by_component
            component.save()


@receiver(post_delete, sender=ProductProduction)
def delete_component_total(sender, instance, **kwargs):
    product_components = instance.product.productcomponent_set.all()
    product = instance.product
    product.total_new -= instance.quantity
    product.save()

    for product_component in product_components:
        total_quantity_by_component = product_component.quantity * instance.quantity
        component = product_component.component
        component.total += total_quantity_by_component
        component.save()


@receiver(post_save, sender=Warehouse)
def add_component_total(sender, instance, created, **kwargs):
    if created:
        print("qo'shildi" * 88)
        component = instance.component
        component.total += instance.quantity
        component.save()


@receiver(pre_save, sender=Warehouse)
def update_component_total_on_edit(sender, instance, **kwargs):
    if instance.pk:
        # Warehouse object is being edited (not a new object)
        old_warehouse = Warehouse.objects.get(pk=instance.pk)
        component = instance.component
        component.total = component.total - old_warehouse.quantity + instance.quantity
        component.save()


@receiver(post_delete, sender=Warehouse)
def update_component_total_on_delete(sender, instance, **kwargs):
    component = instance.component
    component.total -= instance.quantity
    component.save()


@receiver(post_save, sender=SalesEvent)
def sales_create(sender, instance, created, **kwargs):
    if created:
        try:
            product = instance.product
            product.total_new -= instance.quantity_sold
            product.total_sold += instance.quantity_sold
            product.total_sold_price += instance.total_sold_price
            product.save()

        except Exception as e:
            print(e)


@receiver(post_delete, sender=SalesEvent)
def sales_delete(sender, instance, **kwargs):
    try:
        product = instance.product
        product.total_new += instance.quantity_sold
        product.total_sold -= instance.quantity_sold
        product.total_sold_price -= instance.total_sold_price
        product.save()

        sales = instance.sales
        if sales.selling_cut.all().count() == 0:
            sales.delete()
    except Exception as e:
        print(e)


@login_required
def sales_excel_export(request):
    queryset = Sales.objects.all().order_by('-id')

    # Create a new workbook
    workbook = Workbook()
    worksheet = workbook.active

    search_query = request.GET.get('q')
    if search_query:
        queryset = queryset.filter(
            Q(buyer__icontains=search_query) | Q(seller__icontains=search_query))

    filters = request.GET.dict()
    filters.pop('q', None)
    filters.pop('o', None)
    if filters:
        queryset = queryset.filter(**filters)

    # Write headers
    headers = ['Haridor', 'Sotuvchi',
               'Kesilgan mahsulotlar', 'Narx', 'Xodim', 'Sotilgan sana']
    worksheet.append(headers)

    # Write data rows
    for sale in queryset:
        sales_events = sale.selling_cut.all()
        cuts = ", ".join(str(
            sale_event)+f' ({sale_event.single_sold_price} dan)' for sale_event in sales_events)

        total_price = sum(event.total_sold_price for event in sales_events)

        formatted_price = "{:,.1f}".format(total_price)

        row = [
            sale.buyer,
            sale.seller,
            cuts,
            formatted_price,
            sale.user.username if sale.user else '-',
            sale.date.replace(
                tzinfo=None),  # Remove timezone information
        ]
        worksheet.append(row)

    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 100
    worksheet.column_dimensions['D'].width = 50
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 20
    worksheet.column_dimensions['H'].width = 20
    # Set the response headers for file download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=azamat_seh_sales.xlsx'

    # Save workbook to response
    workbook.save(response)

    return response


@login_required
def azamat_production_excel(request):
    queryset = ProductProduction.objects.all().order_by('-id')

    # Create a new workbook
    workbook = Workbook()
    worksheet = workbook.active

    search_query = request.GET.get('q')
    if search_query:
        queryset = queryset.filter(Q(series=search_query))

    filters = request.GET.dict()
    filters.pop('q', None)
    filters.pop('o', None)
    if filters:
        queryset = queryset.filter(**filters)

    # Write headers
    headers = [ 'Produkt', 'Kesilmaganlar soni', 'Xodim', 'Ishlab chiqarilish vaqti']
    worksheet.append(headers)

    # Write data rows
    for production in queryset:
        row = [
            str(production.product),
            production.quantity,
            production.user.username,
            production.date.replace(
                tzinfo=None),  # Remove timezone information
        ]
        worksheet.append(row)

    # Set column width for Arrival Time
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 30

    # Set the response headers for file download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=azamat_production.xlsx'

    # Save workbook to response
    workbook.save(response)

    return response
